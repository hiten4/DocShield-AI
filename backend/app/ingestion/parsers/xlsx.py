import io

from openpyxl import load_workbook

from app.ingestion.parsers import Block
from app.ingestion.parsers.pdf import _table_to_markdown


def parse_xlsx(data: bytes) -> list[Block]:
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    blocks: list[Block] = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([("" if v is None else str(v)).strip() for v in row])
        rows = [r for r in rows if any(r)]
        if not rows:
            continue
        blocks.append(
            Block(
                kind="table",
                text=_table_to_markdown(rows),
                rows=rows,
                section_path=ws.title,
            )
        )
    return blocks
